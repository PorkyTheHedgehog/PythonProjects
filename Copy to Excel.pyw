import openpyxl, heapq
from TextAnalysis import datedict, timedict, persondict, timepdict, worddict
import RunningMacros

RunningMacros.callmacro("ClearAll")

book = openpyxl.load_workbook('Data Analytics.xlsm', keep_vba=True)
sheet = book["Sheet2"]

personL = list(persondict.keys())
timeL = sorted(timedict.keys())
dateL = list(datedict.keys())


for x in range(len(personL)):
	person = personL[x]
	sheet['a'+ str(x+1)] = person
	sheet['b'+ str(x+1)] = persondict[person]

sheet['E1'] = 'Total'

for x in range(len(timeL)):
	time = timeL[x]
	sheet['d'+ str(x+2)] = time
	sheet['e'+ str(x+2)] = timedict[time]

c = 6

for x in timepdict:
	sheet.cell(row=1, column=c).value = x
	for hour in range(0,24):
		sheet.cell(row=hour+2, column=c).value = timepdict[x]["{:02d}".format(hour)]
	c += 1

sheet['J1'] = 'Date'
sheet['K1'] = 'No. Of Messages'

for x in range(len(datedict)):
	date = dateL[x]
	sheet['j' + str(x+2)] = date
	sheet['k' + str(x+2)] = datedict[date]

sheet['M1'] = 'Top 10 words per person:'

c = 14
r = 3
for x in worddict:
	sheet.cell(row=2, column=c).value = x
	for y in heapq.nlargest(10, worddict[x], key=worddict[x].get):
		sheet['M' + str(r)] = y
		if r <= 12:
			sheet.cell(row=r, column=c).value = worddict[x][y]
			sheet.cell(row=r+10, column=c).value = worddict[x][y]
		else:
			sheet.cell(row=r, column=c).value = worddict[x][y]
			sheet.cell(row=r-10, column=c).value = worddict[x][y]
		r += 1
	c += 1


book.save('Data Analytics.xlsm')

RunningMacros.callmacro("DataCleaning")
